# helper script: validator_tasks.py of nyx3ton-project\Few-shots\CV Validator
from __future__ import annotations

import json, os, re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import gradio as gr
from openpyxl import load_workbook
from langchain_core.prompts import ChatPromptTemplate, FewShotChatMessagePromptTemplate

from dictionary_fallback import fallback_extract_requirements_from_text, build_hybrid_requirement_result
from validator_utils import (
    cache_key,
    dedupe_text_list,
    normalize_space,
    normalized_lookup_value,
    parse_boolish,
    remove_diacritics,
    safe_float_value,
    safe_json_loads,
    trim_text,
)
from validator_llm import (
    DEFAULT_FALLBACK_LLM_MODEL_ID,
    DEFAULT_LLM_MODEL_ID,
    LLM_LOAD_MODE,
    SYSTEM_JSON,
    SYSTEM_REQUIREMENT_UTILS_JSON,
    chat_generate_messages,
    lc_messages_to_hf_messages,
)

# -----------------------------------------------------------------------------
# 1. FEW-SHOT EXAMPLES + LANGCHAIN PROMPTS
# -----------------------------------------------------------------------------

def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}

DEFAULT_AUX_LLM_MODEL_ID = os.getenv("AUX_LLM_MODEL_ID", "").strip()
DEFAULT_JOB_SCHEMA_XLSX_PATH = os.getenv("JOB_SCHEMA_XLSX_PATH", str(Path(__file__).resolve().parent / "job_requirement_schema.xlsx"))
DEFAULT_JOB_SCHEMA_SHEET = os.getenv("JOB_SCHEMA_SHEET", "ManualPositions")
DEFAULT_PROMPT_SCHEMA_SHEET = os.getenv("PROMPT_SCHEMA_SHEET", "PromptSchema")
ENABLE_LLM_GENERICNESS = env_bool("ENABLE_LLM_GENERICNESS", True)
ENABLE_LLM_CANONICALIZATION = env_bool("ENABLE_LLM_CANONICALIZATION", True)
GENERIC_HEURISTIC_LOW = float(os.getenv("GENERIC_HEURISTIC_LOW", "0.25"))
GENERIC_HEURISTIC_HIGH = float(os.getenv("GENERIC_HEURISTIC_HIGH", "0.75"))
MIN_LLM_CANONICAL_TEXT_LEN = int(os.getenv("MIN_LLM_CANONICAL_TEXT_LEN", "18"))

_GENERICNESS_CACHE: Dict[str, Dict[str, Any]] = {}
_CANONICAL_CACHE: Dict[str, Dict[str, Any]] = {}

DEFAULT_JOB_REQUIREMENT_SCHEMA = """
{
    "job_title": "",
    "seniority": "junior|medior|senior|unknown",
    "requirements": [
        {
            "id": "R1",
            "text": "konkretna poziadavka",
            "category": "hard_skill|experience|education|language|soft_skill|location|other",
            "priority": "must|nice|unknown",
            "weight": 1.0
        }
    ]
}
""".strip()

CANDIDATE_SUMMARY_SCHEMA = """
{
    "summary": "kratke zhrnutie kandidata bez mena a citlivych atributov",
    "skills": ["..."],
    "experience": ["..."],
    "education": ["..."],
    "languages": ["..."],
    "certifications": ["..."],
    "risks_or_missing_info": ["..."]
}
""".strip()

REQUIREMENT_EVAL_SCHEMA = {
    "requirement_id": "R1",
    "requirement": "konkretna poziadavka",
    "status": "splnene|ciastocne_splnene|nesplnene|nejasne",
    "score": 0,
    "confidence": 0.0,
    "evidence_used": ["kratke citacie/parafrazy dokazov"],
    "explanation": "kratke vysvetlenie v slovencine bez diakritiky",
    "risk_note": "co chyba alebo preco je hodnotenie neiste"
}

JOB_REQ_FEWSHOT = [
    {
        "input": """Hladame Python developera pre backendovy tim. Pozadujeme Python, SQL, Git a REST API. Nutna je prax aspon 2 roky s vyvojom aplikacii. Vyhodou je Docker a Kubernetes. Ocenime anglictinu na urovni B2. Ponukame flexibilny home office, multisport kartu a teambuildingy.""",
        "output": {
            "job_title": "Python developer",
            "seniority": "medior",
            "requirements": [
                {"id": "R1", "text": "Python", "category": "hard_skill", "priority": "must", "weight": 5.0},
                {"id": "R2", "text": "SQL", "category": "hard_skill", "priority": "must", "weight": 4.0},
                {"id": "R3", "text": "Git", "category": "hard_skill", "priority": "must", "weight": 3.0},
                {"id": "R4", "text": "REST API", "category": "hard_skill", "priority": "must", "weight": 3.0},
                {"id": "R5", "text": "aspon 2 roky praxe s vyvojom aplikacii", "category": "experience", "priority": "must", "weight": 4.0},
                {"id": "R6", "text": "Docker", "category": "hard_skill", "priority": "nice", "weight": 2.0},
                {"id": "R7", "text": "Kubernetes", "category": "hard_skill", "priority": "nice", "weight": 2.0},
                {"id": "R8", "text": "anglictina na urovni B2", "category": "language", "priority": "nice", "weight": 2.0}
            ]
        },
    },
    {
        "input": """Pre HR oddelenie hladame HR specialistu. Nutna je skusenost s naborom zamestnancov, komunikacia s kandidátmi a praca s LinkedIn. Vyhodou je SAP SuccessFactors a anglictina. Ocenime samostatnost, zodpovednost a prijemne vystupovanie.""",
        "output": {
            "job_title": "HR specialista",
            "seniority": "unknown",
            "requirements": [
                {"id": "R1", "text": "skusenost s naborom zamestnancov", "category": "experience", "priority": "must", "weight": 5.0},
                {"id": "R2", "text": "komunikacia s kandidatmi", "category": "soft_skill", "priority": "must", "weight": 3.0},
                {"id": "R3", "text": "LinkedIn", "category": "hard_skill", "priority": "must", "weight": 3.0},
                {"id": "R4", "text": "SAP SuccessFactors", "category": "hard_skill", "priority": "nice", "weight": 2.0},
                {"id": "R5", "text": "anglictina", "category": "language", "priority": "nice", "weight": 2.0}
            ]
        },
    },
]

CANDIDATE_FEWSHOT = [
    {
        "input": "Python developer s 3 rocnou praxou. Pracoval s Python, FastAPI, SQL, Docker a Git. Anglictina B2. Studium aplikovanej informatiky.",
        "output": {
            "summary": "Kandidat ma prax vo vyvoji backend aplikacii so zameranim na Python a suvisiace technologie.",
            "skills": ["Python", "FastAPI", "SQL", "Docker", "Git"],
            "experience": ["3 roky praxe vo vyvoji backend aplikacii"],
            "education": ["aplikovana informatika"],
            "languages": ["anglictina B2"],
            "certifications": [],
            "risks_or_missing_info": []
        }
    }
]

REQUIREMENT_EVAL_FEWSHOT = [
    {
        "input": {
            "requirement": {"id": "R1", "text": "Python", "category": "hard_skill", "priority": "must", "weight": 5.0},
            "evidence": ["3 roky komercnej praxe s Python backend vyvojom", "vyvoj REST API vo FastAPI"]
        },
        "output": {
            "requirement_id": "R1",
            "requirement": "Python",
            "status": "splnene",
            "score": 96,
            "confidence": 0.95,
            "evidence_used": ["3 roky komercnej praxe s Python backend vyvojom", "vyvoj REST API vo FastAPI"],
            "explanation": "CV obsahuje priamu a silnu skusenost s Python vyvojom.",
            "risk_note": ""
        }
    },
    {
        "input": {
            "requirement": {"id": "R2", "text": "anglictina na urovni B2", "category": "language", "priority": "must", "weight": 3.0},
            "evidence": ["aktivne pouzivanie anglictiny pri projektoch"]
        },
        "output": {
            "requirement_id": "R2",
            "requirement": "anglictina na urovni B2",
            "status": "ciastocne_splnene",
            "score": 70,
            "confidence": 0.73,
            "evidence_used": ["aktivne pouzivanie anglictiny pri projektoch"],
            "explanation": "CV naznacuje anglictinu, ale explicitna uroven B2 nie je uvedena.",
            "risk_note": "Chyba explicitne uvedena jazykova uroven."
        }
    },
    {
        "input": {
            "requirement": {"id": "R3", "text": "Azure", "category": "hard_skill", "priority": "nice", "weight": 2.0},
            "evidence": []
        },
        "output": {
            "requirement_id": "R3",
            "requirement": "Azure",
            "status": "nesplnene",
            "score": 8,
            "confidence": 0.94,
            "evidence_used": [],
            "explanation": "V dodanych odkazoch z CV nie je dokaz o Azure.",
            "risk_note": "Poziadavku sa nepodarilo podlozit ziadnym dokazom z CV."
        }
    },
]

GENERICNESS_FEWSHOT = [
    {
        "input": "komunikacne schopnosti",
        "output": {
            "is_generic": True,
            "confidence": 0.98,
            "reason": "Ide o vseobecnu soft-skill formulaciu bez konkretneho kontextu.",
            "suggested_focus": "komunikacia so zakaznikom alebo komunikacia v anglictine"
        },
    },
    {
        "input": "anglictina na urovni B2",
        "output": {
            "is_generic": False,
            "confidence": 0.99,
            "reason": "Ide o konkretnu jazykovu poziadavku s explicitnou urovnou.",
            "suggested_focus": ""
        },
    },
    {
        "input": "skusenost s naborom zamestnancov",
        "output": {
            "is_generic": False,
            "confidence": 0.96,
            "reason": "Ide o konkretnu pracovnu skusenost, nie len vseobecnu vlastnost.",
            "suggested_focus": ""
        },
    },
]

CANONICALIZATION_FEWSHOT = [
    {
        "input": "prakticke skusenosti s navrhom a optimalizaciou SQL dotazov",
        "output": {
            "canonical_key": "sql",
            "canonical_text": "SQL",
            "category_hint": "hard_skill",
            "confidence": 0.96,
            "reason": "Hlavnym jadrom poziadavky je SQL."
        },
    },
    {
        "input": "anglictina na urovni B2",
        "output": {
            "canonical_key": "anglictina na urovni b2",
            "canonical_text": "anglictina na urovni B2",
            "category_hint": "language",
            "confidence": 0.99,
            "reason": "Poziadavka je jazykova a explicitne obsahuje uroven B2."
        },
    },
    {
        "input": "schopnost analyzovat business poziadavky a prekladat ich do technickych zadani",
        "output": {
            "canonical_key": "analyza business poziadaviek",
            "canonical_text": "analyza business poziadaviek",
            "category_hint": "other",
            "confidence": 0.87,
            "reason": "Jadrom je analyza a transformacia business poziadaviek do technickej formy."
        },
    },
]


def build_lc_messages_from_fewshot(
    system_text: str,
    examples: List[Dict[str, Any]],
    example_human_template: str,
    final_human_input: str,
) -> List[Dict[str, str]]:
    example_prompt = ChatPromptTemplate.from_messages(
        [
            ("human", example_human_template),
            ("ai", "{output}"),
        ]
    )

    normalized_examples = []
    for ex in examples:
        normalized_examples.append(
            {
                **ex,
                "output": json.dumps(ex["output"], ensure_ascii=False, indent=2),
            }
        )

    few_shot_prompt = FewShotChatMessagePromptTemplate(
        example_prompt=example_prompt,
        examples=normalized_examples,
    )

    final_prompt = ChatPromptTemplate.from_messages(
        [
            ("system", "{system_text}"),
            few_shot_prompt,
            ("human", "{final_human_input}"),
        ]
    )

    prompt_value = final_prompt.format_prompt(
        system_text=system_text,
        final_human_input=final_human_input,
    )

    return lc_messages_to_hf_messages(prompt_value.to_messages())


def load_job_requirement_schema_text(schema_xlsx_path: str) -> Tuple[str, str]:
    path = Path(schema_xlsx_path or "").expanduser()

    if not schema_xlsx_path or not path.exists():
        return DEFAULT_JOB_REQUIREMENT_SCHEMA, "builtin_fallback"

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        if DEFAULT_PROMPT_SCHEMA_SHEET not in wb.sheetnames:
            return DEFAULT_JOB_REQUIREMENT_SCHEMA, "builtin_fallback_missing_prompt_sheet"

        ws = wb[DEFAULT_PROMPT_SCHEMA_SHEET]
        schema_text = ""

        for row in ws.iter_rows(min_row=2, values_only=True):
            key = normalized_lookup_value(row[0] if len(row) > 0 else "")
            value = row[1] if len(row) > 1 else ""
            if key in {"schema_json", "job_requirement_schema", "json_schema"} and str(value or "").strip():
                schema_text = str(value).strip()
                break

        if schema_text:
            return schema_text, f"excel:{path}"

    except Exception:
        pass

    return DEFAULT_JOB_REQUIREMENT_SCHEMA, "builtin_fallback_error"


def normalize_requirement_key_basic(text: str) -> str:
    text = remove_diacritics(normalize_space(text).lower())
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-zA-Z0-9+.#/ ]+", " ", text)

    replacements = {
                    "programovanie v jazyku python": "python",
                    "skusenosti s pythonom": "python",
                    "pokrocila znalost python": "python",
                    "znalost python": "python",
                    "programovanie v jazyku java": "java",
                    "skusenosti s javou": "java",
                    "znalost sql": "sql",
                    "skusenosti so sql": "sql",
                    "anglicky jazyk": "anglictina",
                    "anglictina b2": "anglictina na urovni b2",
                    "english b2": "anglictina na urovni b2",
                    "timova praca": "teamova spolupraca",
                    "praca v time": "teamova spolupraca",
                    "ms office": "microsoft office",
                    "ms excel": "microsoft excel",
                    "microsoft excel": "microsoft excel",
                    "excel": "microsoft excel",
                    "aj": "anglictina",
                    "nj": "nemcina",
                    }

    stop_phrases = [
                    "znalost",
                    "prakticka znalost",
                    "pokrocila znalost",
                    "vyborna znalost",
                    "zakladna znalost",
                    "skusenosti s",
                    "skusenost s",
                    "prax s",
                    "prax v",
                    "schopnost",
                    "schopnost pracovat s",
                    "ovladanie",
                    "aktivna",
                    "vyhodou je",
                    "podmienkou je",
                    "skusenosti v oblasti",
                    ]

    text = replacements.get(text, text)

    for phrase in stop_phrases:
        text = re.sub(rf"\b{re.escape(phrase)}\b", " ", text)

    text = re.sub(r"\s+", " ", text).strip()

    direct_aliases = {
                        "python developer": "python",
                        "python vyvoj": "python",
                        "rest apis": "rest api",
                        "api": "api",
                        "sql databazy": "sql",
                        "english language": "anglictina",
                        "nemecky jazyk": "nemcina",
                        }

    text = direct_aliases.get(text, text)
    return text


def direct_canonical_requirement(text: str) -> Optional[Dict[str, Any]]:
    key = normalize_requirement_key_basic(text)

    mapping = {
                "python": ("python", "Python", "hard_skill"),
                "java": ("java", "Java", "hard_skill"),
                "sql": ("sql", "SQL", "hard_skill"),
                "rest api": ("rest api", "REST API", "hard_skill"),
                "docker": ("docker", "Docker", "hard_skill"),
                "kubernetes": ("kubernetes", "Kubernetes", "hard_skill"),
                "git": ("git", "Git", "hard_skill"),
                "linux": ("linux", "Linux", "hard_skill"),
                "aws": ("aws", "AWS", "hard_skill"),
                "azure": ("azure", "Azure", "hard_skill"),
                "sap successfactors": ("sap successfactors", "SAP SuccessFactors", "hard_skill"),
                "microsoft excel": ("microsoft excel", "Microsoft Excel", "hard_skill"),
                "microsoft office": ("microsoft office", "Microsoft Office", "hard_skill"),
                "anglictina": ("anglictina", "anglictina", "language"),
                "anglictina na urovni b2": ("anglictina na urovni b2", "anglictina na urovni B2", "language"),
                "nemcina": ("nemcina", "nemcina", "language"),
                "teamova spolupraca": ("teamova spolupraca", "teamova spolupraca", "soft_skill"),
                "komunikacne schopnosti": ("komunikacne schopnosti", "komunikacne schopnosti", "soft_skill"),
                "samostatnost": ("samostatnost", "samostatnost", "soft_skill"),
                "zodpovednost": ("zodpovednost", "zodpovednost", "soft_skill"),
                "bratislava": ("bratislava", "Bratislava", "location"),
                "II.stupna": ("magisterske","inzinierske", "magister", "inzinier", "masters"),
                "I.stupna": ("bakalarske", "bakalar", "bachelor"),
                "III.stupna": ("doktorand", "doktorandke", "doctor"),
                "2.stupna": ("magisterske","inzinierske", "magister", "inzinier", "masters"),
                "1.stupna": ("bakalarske", "bakalar", "bachelor"),
                "3.stupna": ("doktorand", "doktorandke", "doctor")
            }

    if key in mapping:
        canonical_key, canonical_text, category_hint = mapping[key]
        return {
                "canonical_key": canonical_key,
                "canonical_text": canonical_text,
                "category_hint": category_hint,
                "confidence": 0.99,
                "reason": "Priama heuristicka zhoda.",
                "source": "heuristic_direct",
                }

    return None


def heuristic_genericness_score(text: str) -> Dict[str, Any]:
    original = normalize_space(text)
    key = normalize_requirement_key_basic(original)
    words = key.split()

    generic_patterns = [
                        "komunikacne schopnosti",
                        "komunikacia",
                        "samostatnost",
                        "zodpovednost",
                        "flexibilita",
                        "odolnost voci stresu",
                        "teamova spolupraca",
                        "teamovy hrac",
                        "proaktivny pristup",
                        "motivacia",
                        "ucit sa nove veci",
                        "prijemne vystupovanie",
                        "spolahlivost",
                        ]

    concrete_patterns = [
                        "python",
                        "java",
                        "sql",
                        "excel",
                        "microsoft excel",
                        "docker",
                        "kubernetes",
                        "linux",
                        "aws",
                        "azure",
                        "sap",
                        "rest api",
                        "api",
                        "git",
                        "b1",
                        "b2",
                        "c1",
                        "c2",
                        "bratislava",
                        "praxe",
                        "rok",
                        "roky",
                        "certifikat",
                        "vzdelanie",
                        "bakalar",
                        "magister",
                        ]

    score = 0.50
    reasons = []

    if any(p == key or p in key for p in generic_patterns):
        score += 0.32
        reasons.append("Obsahuje vseobecnu soft-skill formulaciu.")

    if len(words) <= 2 and not re.search(r"\d", key):
        score += 0.12
        reasons.append("Poziadavka je velmi kratka a bez meratelneho detailu.")

    if re.search(r"\b(aspon|min\.?|minimalne)\b", key):
        score -= 0.18
        reasons.append("Obsahuje minimalnu alebo meratelnu podmienku.")

    if re.search(r"\b(a1|a2|b1|b2|c1|c2)\b", key):
        score -= 0.24
        reasons.append("Obsahuje explicitnu jazykovu uroven.")

    if re.search(r"\b\d+\b", key):
        score -= 0.20
        reasons.append("Obsahuje cislo alebo kvantifikaciu.")

    if any(p in key for p in concrete_patterns):
        score -= 0.30
        reasons.append("Obsahuje konkretny nastroj, technologiu alebo meratelny signal.")

    if len(words) >= 6:
        score -= 0.08
        reasons.append("Poziadavka je opisnejsia a konkretnejsia.")

    score = max(0.0, min(1.0, score))
    confidence = 0.55 + abs(score - 0.5)
    confidence = max(0.0, min(0.99, confidence))

    return {
        "is_generic": score >= 0.50,
        "confidence": round(confidence, 4),
        "heuristic_score": round(score, 4),
        "reason": " ".join(reasons) if reasons else "Bez silneho heuristickeho signalu.",
        "source": "heuristic",
    }


def resolve_assist_model_id(primary_model_id: str, aux_model_id: Optional[str] = None) -> str:
    return (aux_model_id or DEFAULT_AUX_LLM_MODEL_ID or primary_model_id).strip()


def llm_classify_requirement_genericness(
    text: str,
    model_id: str,
    load_mode: str,
    fallback_model_id: str,
    aux_model_id: Optional[str] = None,
) -> Dict[str, Any]:
    assist_model_id = resolve_assist_model_id(model_id, aux_model_id)
    ck = cache_key("genericness", text, assist_model_id)
    if ck in _GENERICNESS_CACHE:
        return _GENERICNESS_CACHE[ck]

    final_user = (
                    "Posud, ci je poziadavka na kandidata prilis vseobecna alebo naopak dost konkretna.\n\n"
                    "Vrat presne tento JSON tvar:\n"
                    "{\n"
                    '  "is_generic": true,\n'
                    '  "confidence": 0.0,\n'
                    '  "reason": "kratke vysvetlenie",\n'
                    '  "suggested_focus": "ak je to vhodne, kratke konkretne jadro poziadavky, inak prazdny string"\n'
                    "}\n\n"
                    "Pravidla:\n"
                    "- true pouzi pri vseobecnych soft-skill formulaciach bez konkretneho obsahu\n"
                    "- false pouzi pri technologii, jazyku, meratelnej poziadavke, nastroji, praxi alebo lokalite\n"
                    "- vrat iba JSON\n\n"
                    "POZIADAVKA:\n"
                    + normalize_space(text)
                )

    messages = build_lc_messages_from_fewshot(
                                                system_text=SYSTEM_REQUIREMENT_UTILS_JSON,
                                                examples=GENERICNESS_FEWSHOT,
                                                example_human_template="Posud generickost poziadavky.\n\nPOZIADAVKA:\n{input}",
                                                final_human_input=final_user,
                                            )

    raw = chat_generate_messages(
                                messages,
                                assist_model_id,
                                load_mode,
                                fallback_model_id,
                                max_new_tokens=220,
                                do_sample=False,
                                )

    data = safe_json_loads(
        raw,
        fallback={
                    "is_generic": False,
                    "confidence": 0.0,
                    "reason": "LLM klasifikacia zlyhala.",
                    "suggested_focus": "",
                },
    )

    if not isinstance(data, dict):
        data = {
                    "is_generic": False,
                    "confidence": 0.0,
                    "reason": "LLM klasifikacia zlyhala.",
                    "suggested_focus": "",
                }

    result = {
                "is_generic": bool(data.get("is_generic", False)),
                "confidence": max(0.0, min(1.0, float(data.get("confidence", 0.0) or 0.0))),
                "heuristic_score": None,
                "reason": normalize_space(str(data.get("reason") or "")),
                "suggested_focus": normalize_space(str(data.get("suggested_focus") or "")),
                "source": "llm_genericness",
                }

    _GENERICNESS_CACHE[ck] = result
    return result


def score_requirement_genericness(
    text: str,
    model_id: Optional[str] = None,
    load_mode: Optional[str] = None,
    fallback_model_id: Optional[str] = None,
    aux_model_id: Optional[str] = None,
    allow_llm: bool = True,
) -> Dict[str, Any]:
    heuristic = heuristic_genericness_score(text)

    if heuristic["heuristic_score"] <= GENERIC_HEURISTIC_LOW:
        return {
                **heuristic,
                "is_generic": False,
                "source": "heuristic_strong_non_generic",
                }

    if heuristic["heuristic_score"] >= GENERIC_HEURISTIC_HIGH:
        return {
                **heuristic,
                "is_generic": True,
                "source": "heuristic_strong_generic",
                }

    if not allow_llm or not ENABLE_LLM_GENERICNESS:
        return heuristic

    return llm_classify_requirement_genericness(
                                                text=text,
                                                model_id=model_id or DEFAULT_LLM_MODEL_ID,
                                                load_mode=load_mode or LLM_LOAD_MODE,
                                                fallback_model_id=fallback_model_id or DEFAULT_FALLBACK_LLM_MODEL_ID,
                                                aux_model_id=aux_model_id,
                                                )


def is_generic_requirement_text(
    text: str,
    model_id: Optional[str] = None,
    load_mode: Optional[str] = None,
    fallback_model_id: Optional[str] = None,
    aux_model_id: Optional[str] = None,
    allow_llm: bool = True,
) -> bool:
    data = score_requirement_genericness(
        text=text,
        model_id=model_id,
        load_mode=load_mode,
        fallback_model_id=fallback_model_id,
        aux_model_id=aux_model_id,
        allow_llm=allow_llm,
    )
    return bool(data.get("is_generic", False))


def llm_canonicalize_requirement(
    text: str,
    model_id: str,
    load_mode: str,
    fallback_model_id: str,
    aux_model_id: Optional[str] = None,
) -> Dict[str, Any]:
    assist_model_id = resolve_assist_model_id(model_id, aux_model_id)
    ck = cache_key("canonical", text, assist_model_id)
    if ck in _CANONICAL_CACHE:
        return _CANONICAL_CACHE[ck]

    final_user = (
                    "Znormalizuj jednu poziadavku na kandidata do kanonickeho kluca vhodneho na deduplikaciu.\n\n"
                    "Vrat presne tento JSON tvar:\n"
                    "{\n"
                    '  "canonical_key": "kratky kluc malymi pismenami",\n'
                    '  "canonical_text": "kratka citatelna verzia",\n'
                    '  "category_hint": "hard_skill|experience|education|language|soft_skill|location|other|unknown",\n'
                    '  "confidence": 0.0,\n'
                    '  "reason": "kratke vysvetlenie"\n'
                    "}\n\n"
                    "Pravidla:\n"
                    "- canonical_key musi byt kratky, stabilny a bez zbytocnych slov\n"
                    "- zachovaj hlavny skill, jazyk, lokalitu alebo jadro poziadavky\n"
                    "- nepis vetu, ale jadro poziadavky\n"
                    "- vrat iba JSON\n\n"
                    "POZIADAVKA:\n"
                    + normalize_space(text)
                )

    messages = build_lc_messages_from_fewshot(
                                                system_text=SYSTEM_REQUIREMENT_UTILS_JSON,
                                                examples=CANONICALIZATION_FEWSHOT,
                                                example_human_template="Znormalizuj poziadavku do kanonickej formy.\n\nPOZIADAVKA:\n{input}",
                                                final_human_input=final_user,
                                            )

    raw = chat_generate_messages(
                                    messages,
                                    assist_model_id,
                                    load_mode,
                                    fallback_model_id,
                                    max_new_tokens=260,
                                    do_sample=False,
                                )

    data = safe_json_loads(
                            raw,
                            fallback={
                            "canonical_key": normalize_requirement_key_basic(text),
                            "canonical_text": normalize_space(text),
                            "category_hint": "unknown",
                            "confidence": 0.0,
                            "reason": "LLM kanonizacia zlyhala.",
                            },
                            )

    if not isinstance(data, dict):
        data = {
                "canonical_key": normalize_requirement_key_basic(text),
                "canonical_text": normalize_space(text),
                "category_hint": "unknown",
                "confidence": 0.0,
                "reason": "LLM kanonizacia zlyhala.",
                }

    canonical_key = normalize_requirement_key_basic(str(data.get("canonical_key") or text))
    canonical_text = normalize_space(str(data.get("canonical_text") or text))
    category_hint = str(data.get("category_hint") or "unknown")
    if category_hint not in {"hard_skill", "experience", "education", "language", "soft_skill", "location", "other", "unknown"}:
        category_hint = "unknown"

    result = {
                "canonical_key": canonical_key or normalize_requirement_key_basic(text),
                "canonical_text": canonical_text or normalize_space(text),
                "category_hint": category_hint,
                "confidence": max(0.0, min(1.0, float(data.get("confidence", 0.0) or 0.0))),
                "reason": normalize_space(str(data.get("reason") or "")),
                "source": "llm_canonicalizer",
            }

    _CANONICAL_CACHE[ck] = result
    return result


def hybrid_normalize_requirement_key(
    text: str,
    model_id: Optional[str] = None,
    load_mode: Optional[str] = None,
    fallback_model_id: Optional[str] = None,
    aux_model_id: Optional[str] = None,
    allow_llm: bool = True,
) -> Dict[str, Any]:
    text = normalize_space(text)
    direct = direct_canonical_requirement(text)
    if direct is not None:
        return direct

    basic_key = normalize_requirement_key_basic(text)
    heuristic_result = {
                            "canonical_key": basic_key,
                            "canonical_text": text,
                            "category_hint": "unknown",
                            "confidence": 0.68 if len(basic_key.split()) <= 3 else 0.52,
                            "reason": "Heuristicka kanonizacia bez LLM.",
                            "source": "heuristic",
                        }

    if not allow_llm or not ENABLE_LLM_CANONICALIZATION:
        return heuristic_result

    if len(text) < MIN_LLM_CANONICAL_TEXT_LEN and len(basic_key.split()) <= 3:
        return heuristic_result

    llm_result = llm_canonicalize_requirement(
                                                text=text,
                                                model_id=model_id or DEFAULT_LLM_MODEL_ID,
                                                load_mode=load_mode or LLM_LOAD_MODE,
                                                fallback_model_id=fallback_model_id or DEFAULT_FALLBACK_LLM_MODEL_ID,
                                                aux_model_id=aux_model_id,
                                            )

    if not llm_result.get("canonical_key"):
        return heuristic_result

    return llm_result


def normalize_requirement_key(
    text: str,
    model_id: Optional[str] = None,
    load_mode: Optional[str] = None,
    fallback_model_id: Optional[str] = None,
    aux_model_id: Optional[str] = None,
    allow_llm: bool = True,
) -> str:
    data = hybrid_normalize_requirement_key(
        text=text,
        model_id=model_id or DEFAULT_LLM_MODEL_ID,
        load_mode=load_mode or LLM_LOAD_MODE,
        fallback_model_id=fallback_model_id or DEFAULT_FALLBACK_LLM_MODEL_ID,
        aux_model_id=aux_model_id,
        allow_llm=allow_llm,
    )
    return str(data.get("canonical_key") or normalize_requirement_key_basic(text))


def clean_requirements(
    reqs: List[Any],
    max_requirements: int,
    model_id: Optional[str] = None,
    load_mode: Optional[str] = None,
    fallback_model_id: Optional[str] = None,
    aux_model_id: Optional[str] = None,
    allow_llm: bool = True,
) -> List[Dict[str, Any]]:
    clean_reqs = []
    seen_keys = set()

    for i, r in enumerate(reqs[: max_requirements * 3], start=1):
        if not isinstance(r, dict):
            continue

        text = normalize_space(str(r.get("text", "")))
        if not text:
            continue

        canonical = hybrid_normalize_requirement_key(
                                                        text=text,
                                                        model_id=model_id,
                                                        load_mode=load_mode,
                                                        fallback_model_id=fallback_model_id,
                                                        aux_model_id=aux_model_id,
                                                        allow_llm=allow_llm,
                                                    )

        key = str(canonical.get("canonical_key") or normalize_requirement_key_basic(text))
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)

        try:
            weight = float(r.get("weight", 1.0))
        except Exception:
            weight = 1.0

        category = str(r.get("category") or "other")
        if category not in {"hard_skill", "experience", "education", "language", "soft_skill", "location", "other"}:
            category = "other"

        category_hint = str(canonical.get("category_hint") or "unknown")
        if category == "other" and category_hint in {"hard_skill", "experience", "education", "language", "soft_skill", "location"}:
            category = category_hint

        priority = str(r.get("priority") or "unknown")
        if priority not in {"must", "nice", "unknown"}:
            priority = "unknown"

        generic_meta = score_requirement_genericness(
                                                        text=text,
                                                        model_id=model_id,
                                                        load_mode=load_mode,
                                                        fallback_model_id=fallback_model_id,
                                                        aux_model_id=aux_model_id,
                                                        allow_llm=allow_llm,
                                                    )

        clean_reqs.append(
                            {
                                "id": str(r.get("id") or f"R{i}"),
                                "text": text,
                                "category": category,
                                "priority": priority,
                                "weight": max(0.5, min(5.0, weight)),
                                "canonical_key": key,
                                "generic_flag": bool(generic_meta.get("is_generic", False)),
                            }
                        )

        if len(clean_reqs) >= max_requirements:
            break

    for idx, item in enumerate(clean_reqs, start=1):
        item["id"] = f"R{idx}"

    return clean_reqs


def is_weak_requirement_output(
    data: Dict[str, Any],
    max_requirements: int,
    model_id: Optional[str] = None,
    load_mode: Optional[str] = None,
    fallback_model_id: Optional[str] = None,
    aux_model_id: Optional[str] = None,
    allow_llm: bool = True,
) -> bool:
    reqs = data.get("requirements") or []
    if not isinstance(reqs, list):
        return True

    if len(reqs) < min(3, max_requirements):
        return True

    generic_count = 0
    hard_like_count = 0
    seen = set()
    dup_count = 0

    for req in reqs:
        if not isinstance(req, dict):
            continue

        text = normalize_space(str(req.get("text", "")))
        key = str(req.get("canonical_key") or normalize_requirement_key(
                                                                            text,
                                                                            model_id=model_id,
                                                                            load_mode=load_mode,
                                                                            fallback_model_id=fallback_model_id,
                                                                            aux_model_id=aux_model_id,
                                                                            allow_llm=allow_llm,
                                                                        ))

        if not key:
            continue

        if key in seen:
            dup_count += 1
        seen.add(key)

        if bool(req.get("generic_flag")) or is_generic_requirement_text(
                                                                            text,
                                                                            model_id=model_id,
                                                                            load_mode=load_mode,
                                                                            fallback_model_id=fallback_model_id,
                                                                            aux_model_id=aux_model_id,
                                                                            allow_llm=allow_llm,
        ):
            generic_count += 1

        category = str(req.get("category") or "other")
        if category in {"hard_skill", "experience", "education", "language"}:
            hard_like_count += 1

    if dup_count >= 2:
        return True
    if generic_count >= max(2, len(reqs) // 2):
        return True
    if hard_like_count == 0:
        return True

    return False


def load_manual_job_requirements_from_excel(
    position_query: str,
    schema_xlsx_path: str,
    max_requirements: int,
    model_id: str,
    load_mode: str,
    fallback_model_id: str,
    aux_model_id: str,
) -> Dict[str, Any]:
    path = Path(schema_xlsx_path or "").expanduser()

    if not position_query or not position_query.strip():
        raise gr.Error("Pre manualnu poziciu zo schema XLSX zadaj nazov pozicie alebo position_key.")

    if not schema_xlsx_path or not path.exists():
        raise gr.Error("Schema XLSX subor neexistuje. Skontroluj JOB_SCHEMA_XLSX_PATH alebo pole v UI.")

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise gr.Error(f"Nepodarilo sa otvorit schema XLSX: {exc}") from exc

    if DEFAULT_JOB_SCHEMA_SHEET not in wb.sheetnames:
        raise gr.Error(f"V schema XLSX chyba harok '{DEFAULT_JOB_SCHEMA_SHEET}'.")

    ws = wb[DEFAULT_JOB_SCHEMA_SHEET]
    rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    headers_row = next(rows_iter, None)
    if not headers_row:
        raise gr.Error("Harok ManualPositions nema hlavicku.")

    normalized_headers = [normalized_lookup_value(h) for h in headers_row]

    header_aliases = {
                        "position_key": {"position_key", "position id", "position_id", "role_key", "job_key"},
                        "job_title": {"job_title", "job title", "pozicia", "position_name", "role_name", "nazov_pozicie"},
                        "seniority": {"seniority", "seniorita", "level"},
                        "requirement_id": {"requirement_id", "requirement id", "req_id", "id"},
                        "requirement_text": {"requirement_text", "requirement text", "text", "poziadavka", "requirement"},
                        "category": {"category", "kategoria"},
                        "priority": {"priority", "priorita"},
                        "weight": {"weight", "vaha"},
                        "active": {"active", "aktivne", "enabled", "is_active"},
                        "note": {"note", "notes", "poznamka"},
                    }

    header_positions: Dict[str, int] = {}
    for canonical_name, variants in header_aliases.items():
        for idx, header_name in enumerate(normalized_headers):
            if header_name in variants:
                header_positions[canonical_name] = idx
                break

    required_headers = {"job_title", "requirement_text"}
    missing = [x for x in required_headers if x not in header_positions]
    if missing:
        raise gr.Error(
            "Harok ManualPositions nema povinne stlpce: " + ", ".join(missing) + "."
        )

    grouped_rows: Dict[str, List[Dict[str, Any]]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue

        def get_value(name: str, default: Any = "") -> Any:
            idx = header_positions.get(name)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        requirement_text = normalize_space(str(get_value("requirement_text", "")))
        if not requirement_text:
            continue

        active = parse_boolish(get_value("active", True), default=True)
        if not active:
            continue

        position_key = normalize_space(str(get_value("position_key", "")))
        job_title = normalize_space(str(get_value("job_title", "")))
        seniority = normalize_space(str(get_value("seniority", "unknown"))) or "unknown"

        logical_group = position_key or job_title
        if not logical_group:
            continue

        grouped_rows.setdefault(logical_group, []).append(
            {
                "position_key": position_key,
                "job_title": job_title or logical_group,
                "seniority": seniority,
                "requirement_id": normalize_space(str(get_value("requirement_id", ""))),
                "requirement_text": requirement_text,
                "category": normalize_space(str(get_value("category", "other"))) or "other",
                "priority": normalize_space(str(get_value("priority", "unknown"))) or "unknown",
                "weight": safe_float_value(get_value("weight", 1.0), 1.0),
                "note": normalize_space(str(get_value("note", ""))),
            }
        )

    if not grouped_rows:
        raise gr.Error("Schema XLSX neobsahuje ziadne aktivne manualne poziadavky.")

    query_norm = normalized_lookup_value(position_query)
    best_group_name = ""
    best_score = -1

    for group_name, group_rows in grouped_rows.items():
        sample = group_rows[0]
        title_norm = normalized_lookup_value(sample.get("job_title", ""))
        key_norm = normalized_lookup_value(sample.get("position_key", ""))
        group_norm = normalized_lookup_value(group_name)

        score = 0
        if query_norm and query_norm == key_norm:
            score = 100
        elif query_norm and query_norm == title_norm:
            score = 98
        elif query_norm and query_norm == group_norm:
            score = 96
        elif query_norm and key_norm and query_norm in key_norm:
            score = 90
        elif query_norm and title_norm and query_norm in title_norm:
            score = 88
        elif query_norm and group_norm and query_norm in group_norm:
            score = 86
        elif query_norm and key_norm and key_norm in query_norm:
            score = 80
        elif query_norm and title_norm and title_norm in query_norm:
            score = 78

        if score > best_score:
            best_group_name = group_name
            best_score = score

    if best_score < 0 or best_group_name not in grouped_rows:
        raise gr.Error(f"V schema XLSX sa nenasla manualna pozicia pre '{position_query}'.")

    matched_rows = grouped_rows[best_group_name]
    sample = matched_rows[0]

    job_data = {
                "job_title": sample.get("job_title") or best_group_name,
                "seniority": sample.get("seniority") or "unknown",
                "requirements": [],
                "_source": "external_excel_schema",
                "_meta": {
                            "schema_path": str(path.resolve()),
                            "schema_sheet": DEFAULT_JOB_SCHEMA_SHEET,
                            "manual_position_query": position_query,
                            "matched_group": best_group_name,
                            "match_score": best_score,
                            "prompt_mode": "manual_excel_schema",
                        },
                }

    raw_requirements = []
    for i, item in enumerate(matched_rows, start=1):
        raw_requirements.append(
            {
                "id": item.get("requirement_id") or f"R{i}",
                "text": item.get("requirement_text", ""),
                "category": item.get("category", "other"),
                "priority": item.get("priority", "unknown"),
                "weight": item.get("weight", 1.0),
                "note": item.get("note", ""),
            }
        )

    job_data["requirements"] = clean_requirements(
                                                    raw_requirements,
                                                    max_requirements,
                                                    model_id=model_id,
                                                    load_mode=load_mode,
                                                    fallback_model_id=fallback_model_id,
                                                    aux_model_id=aux_model_id,
                                                    allow_llm=True,
                                                    )

    return job_data

# -----------------------------------------------------------------------------
# 9. LLM PROMPTS / TASKS
# -----------------------------------------------------------------------------

def extract_job_requirements(
    job_text: str,
    job_requirement_schema_text: str,
    model_id: str,
    load_mode: str,
    fallback_model_id: str,
    aux_model_id: str,
    max_requirements: int,
) -> Dict[str, Any]:
    final_user = (
                    "Z pracovnej ponuky extrahuj atomicke poziadavky na kandidata.\n\n"
                    "Vrat presne tento JSON tvar:\n"
                    + job_requirement_schema_text
                    + "\n\nPravidla:\n"
                    f"- maximalne {max_requirements} najdolezitejsich poziadaviek\n"
                    "- must-have poziadavky daj weight 3 az 5\n"
                    "- nice-to-have poziadavky daj weight 1 az 2\n"
                    "- nerozbijaj jednu poziadavku na duplicitne varianty\n"
                    "- ignoruj benefity firmy, marketing a pravne formulacie\n"
                    "- ak nevies urcit job_title alebo seniority, pouzi hodnotu unknown\n"
                    "- vrat iba JSON, bez komentara\n\n"
                    "PRACOVNA PONUKA:\n"
                    + trim_text(job_text, 22000)
                    )

    messages = build_lc_messages_from_fewshot(
                                                system_text=SYSTEM_JSON,
                                                examples=JOB_REQ_FEWSHOT,
                                                example_human_template="Z pracovnej ponuky extrahuj atomicke poziadavky na kandidata.\n\nPRACOVNA PONUKA:\n{input}",
                                                final_human_input=final_user,
                                            )

    raw = chat_generate_messages(
                                    messages,
                                    model_id,
                                    load_mode,
                                    fallback_model_id,
                                    max_new_tokens=1100,
                                    do_sample=False,
                                )

    print("\n   RAW JOB REQUIREMENTS OUTPUT ")
    print(raw)
    print(" END RAW JOB REQUIREMENTS OUTPUT \n")

    llm_data = safe_json_loads(raw, fallback={"job_title": "unknown", "seniority": "unknown", "requirements": []})

    if isinstance(llm_data, list):
        llm_data = {"job_title": "unknown", "seniority": "unknown", "requirements": llm_data}
    if not isinstance(llm_data, dict):
        llm_data = {"job_title": "unknown", "seniority": "unknown", "requirements": []}

    llm_data["requirements"] = clean_requirements(
                                                    llm_data.get("requirements") or [],
                                                    max_requirements,
                                                    model_id=model_id,
                                                    load_mode=load_mode,
                                                    fallback_model_id=fallback_model_id,
                                                    aux_model_id=aux_model_id,
                                                    allow_llm=True,
                                                    )
    llm_data["_source"] = "llm_few_shot"

    weak_llm = is_weak_requirement_output(
                                            llm_data,
                                            max_requirements=max_requirements,
                                            model_id=model_id,
                                            load_mode=load_mode,
                                            fallback_model_id=fallback_model_id,
                                            aux_model_id=aux_model_id,
                                            allow_llm=True,
                                        )
    fallback_data = fallback_extract_requirements_from_text(job_text, max_requirements)

    hybrid_data = build_hybrid_requirement_result(
                                                    llm_data=llm_data,
                                                    fallback_data=fallback_data,
                                                    max_requirements=max_requirements,
                                                )

    if not isinstance(hybrid_data, dict):
        hybrid_data = llm_data

    hybrid_data.setdefault("_meta", {})
    if isinstance(hybrid_data["_meta"], dict):
        hybrid_data["_meta"]["weak_llm"] = weak_llm
        hybrid_data["_meta"]["prompt_mode"] = "few_shot_langchain"

    hybrid_data["requirements"] = clean_requirements(
                                                        hybrid_data.get("requirements") or [],
                                                        max_requirements,
                                                        model_id=model_id,
                                                        load_mode=load_mode,
                                                        fallback_model_id=fallback_model_id,
                                                        aux_model_id=aux_model_id,
                                                        allow_llm=True,
                                                    )

    print("\n   REQUIREMENT EXTRACTION SUMMARY  ")
    print(json.dumps(hybrid_data.get("_meta", {}), ensure_ascii=False, indent=2))
    print(f"Source: {hybrid_data.get('_source', 'unknown')}")
    print(" END REQUIREMENT EXTRACTION SUMMARY  \n")

    return hybrid_data


def extract_candidate_summary(
    cv_text: str,
    model_id: str,
    load_mode: str,
    fallback_model_id: str,
) -> Dict[str, Any]:
    final_user = (
                    "Zo zivotopisu extrahuj anonymizovany profil kandidata.\n\n"
                    "Vrat presne tento JSON tvar:\n"
                    + CANDIDATE_SUMMARY_SCHEMA
                    + "\n\nPravidla:\n"
                    "- nepouzivaj meno, adresu, vek, pohlavie, rodinny stav ani fotku\n"
                    "- uvadzaj len veci, ktore su v CV\n"
                    "- vrat iba JSON, bez komentara\n\n"
                    "CV TEXT:\n"
                    + trim_text(cv_text, 22000)
                    )

    messages = build_lc_messages_from_fewshot(
                                                system_text=SYSTEM_JSON,
                                                examples=CANDIDATE_FEWSHOT,
                                                example_human_template="Zo zivotopisu extrahuj anonymizovany profil kandidata.\n\nCV TEXT:\n{input}",
                                                final_human_input=final_user,
                                            )

    raw = chat_generate_messages(
                                    messages,
                                    model_id,
                                    load_mode,
                                    fallback_model_id,
                                    max_new_tokens=900,
                                    do_sample=False,
                                )

    data = safe_json_loads(raw, fallback={"summary": "Nepodarilo sa spolahlivo extrahovat profil.", "skills": []})
    if not isinstance(data, dict):
        data = {"summary": "Nepodarilo sa spolahlivo extrahovat profil.", "skills": []}

    for key in ["skills", "experience", "education", "languages", "certifications", "risks_or_missing_info"]:
        vals = data.get(key)
        data[key] = dedupe_text_list(vals if isinstance(vals, list) else [], limit=12)

    return data


def evaluate_one_requirement(
    requirement: Dict[str, Any],
    evidence: List[str],
    model_id: str,
    load_mode: str,
    fallback_model_id: str,
) -> Dict[str, Any]:
    final_user = (
                    "Vyhodnot, ci kandidat splna jednu poziadavku z pracovneho inzeratu.\n"
                    "Pouzi iba dodane odkazy z CV. Ak dokaz chyba, nehadaj.\n\n"
                    "Vrat presne tento JSON tvar:\n"
                    + json.dumps(REQUIREMENT_EVAL_SCHEMA, ensure_ascii=False, indent=2)
                    + "\n\nSkorovanie:\n"
                    "- 90-100 = jasne splnene\n"
                    "- 60-89 = skor splnene alebo ciastocne\n"
                    "- 30-59 = slabe/nepriame odkazy\n"
                    "- 0-29 = nesplnene alebo chyba dokaz\n\n"
                    "POZIADAVKA:\n"
                    + json.dumps(requirement, ensure_ascii=False, indent=2)
                    + "\n\nODKAZY Z CV:\n"
                    + json.dumps(evidence, ensure_ascii=False, indent=2)
                    )

    eval_examples = []
    for ex in REQUIREMENT_EVAL_FEWSHOT:
        eval_examples.append(
            {
                "input": (
                    "POZIADAVKA:\n"
                    + json.dumps(ex["input"]["requirement"], ensure_ascii=False, indent=2)
                    + "\n\nODKAZY Z CV:\n"
                    + json.dumps(ex["input"]["evidence"], ensure_ascii=False, indent=2)
                ),
                "output": ex["output"],
            }
        )

    messages = build_lc_messages_from_fewshot(
                                                system_text=SYSTEM_JSON,
                                                examples=eval_examples,
                                                example_human_template="Vyhodnot, ci kandidat splna jednu poziadavku z pracovneho inzeratu.\n\n{input}",
                                                final_human_input=final_user,
                                            )

    raw = chat_generate_messages(
                                    messages,
                                    model_id,
                                    load_mode,
                                    fallback_model_id,
                                    max_new_tokens=700,
                                    do_sample=False,
                                )

    data = safe_json_loads(raw, fallback={})
    if not isinstance(data, dict):
        data = {}

    status = str(data.get("status") or "nejasne")
    if status not in {"splnene", "ciastocne_splnene", "nesplnene", "nejasne"}:
        status = "nejasne"

    try:
        score = float(data.get("score", 0))
    except Exception:
        score = 0.0

    try:
        confidence = float(data.get("confidence", 0))
    except Exception:
        confidence = 0.0

    evidence_used = data.get("evidence_used") if isinstance(data.get("evidence_used"), list) else evidence[:2]
    evidence_used = dedupe_text_list([str(x) for x in evidence_used], limit=3)

    return {
            "requirement_id": str(data.get("requirement_id") or requirement.get("id", "")),
            "requirement": str(data.get("requirement") or requirement.get("text", "")),
            "category": requirement.get("category", "other"),
            "priority": requirement.get("priority", "unknown"),
            "weight": requirement.get("weight", 1.0),
            "status": status,
            "score": max(0.0, min(100.0, score)),
            "confidence": max(0.0, min(1.0, confidence)),
            "evidence_used": evidence_used,
            "explanation": str(data.get("explanation") or "Bez vysvetlenia."),
            "risk_note": str(data.get("risk_note") or ""),
            }
