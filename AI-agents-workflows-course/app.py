from __future__ import annotations

import json, os, re, smtplib, ssl, time, requests, truststore
truststore.inject_into_ssl()
from dataclasses import asdict, dataclass
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import gradio as gr
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OPEN_METEO_GEOCODING = "https://geocoding-api.open-meteo.com/v1/search"
OPEN_METEO_FORECAST = "https://api.open-meteo.com/v1/forecast"

OPENVINO_MODEL_DIR = os.getenv("OPENVINO_MODEL_DIR", "").strip()
OPENVINO_MODEL_ID = os.getenv("OPENVINO_MODEL_ID", "unsloth/Qwen3-4B-Instruct-2507").strip()
OPENVINO_DEVICE = os.getenv("OPENVINO_DEVICE", "CPU").strip() or "CPU"
OPENVINO_WEIGHT_FORMAT = os.getenv("OPENVINO_WEIGHT_FORMAT", "int8").strip().lower()
OV_CACHE_DIR = BASE_DIR / "ov_cache"

CUDA_MODEL_ID = os.getenv("CUDA_MODEL_ID", "").strip()
CUDA_DEVICE = os.getenv("CUDA_DEVICE", "cuda:0").strip() or "cuda:0"
CUDA_QUANTIZATION = (os.getenv("CUDA_QUANTIZATION", "auto").strip().lower() or "auto")

REASONER_BACKEND = (os.getenv("REASONER_BACKEND", "auto").strip().lower() or "auto")

SMTP_HOST = os.getenv("SMTP_HOST", "").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "587") or 587)
SMTP_USER = os.getenv("SMTP_USER", "").strip()
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "").strip()
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USER).strip()

SMTP_SECURITY = (os.getenv("SMTP_SECURITY", "starttls").strip().lower() or "starttls")
PREDEFINED_RECIPIENT = os.getenv("PREDEFINED_RECIPIENT", "").strip()

CUSTOM_CSS = """
.gradio-container { max-width: 1500px !important; }
.agent-box textarea { font-size: 1rem !important; line-height: 1.45 !important; }
"""

WEATHER_CODES = {
                0: "Clear sky",
                1: "Mainly clear",
                2: "Partly cloudy",
                3: "Overcast",
                45: "Fog",
                48: "Depositing rime fog",
                51: "Light drizzle",
                53: "Moderate drizzle",
                55: "Dense drizzle",
                56: "Light freezing drizzle",
                57: "Dense freezing drizzle",
                61: "Slight rain",
                63: "Moderate rain",
                65: "Heavy rain",
                66: "Light freezing rain",
                67: "Heavy freezing rain",
                71: "Slight snow",
                73: "Moderate snow",
                75: "Heavy snow",
                77: "Snow grains",
                80: "Slight rain showers",
                81: "Moderate rain showers",
                82: "Violent rain showers",
                85: "Slight snow showers",
                86: "Heavy snow showers",
                95: "Thunderstorm",
                96: "Thunderstorm with slight hail",
                96: "Thunderstorm with heavy hail"
                }

@dataclass
class AgentEvent:
                agent: str
                step: str
                status: str
                detail: str
                elapsed_sec: float
class Trace:
    def __init__(self) -> None:
        self.events: list[AgentEvent] = []

    def add(self, agent: str, step: str, status: str, detail: str, start: float) -> None:
        self.events.append(AgentEvent(agent=agent,step=step,status=status,detail=detail,elapsed_sec=round(time.perf_counter() - start, 3)))

    def dataframe(self) -> pd.DataFrame:
        return pd.DataFrame([asdict(event) for event in self.events])


def detect_cuda() -> tuple[bool, str]:

    try:
        import torch  # type: ignore[import-not-found]
    except Exception as exc:  # torch not installed
        return False, f"PyTorch not available: {exc}"

    try:
        if not torch.cuda.is_available():
            return False, "PyTorch is installed but no CUDA device is available."
        name = torch.cuda.get_device_name(0)
        count = torch.cuda.device_count()
        return True, f"CUDA available: {count} device(s), primary GPU '{name}'."
    except Exception as exc:
        return False, f"CUDA detection failed: {exc}"

def resolve_model_path(model_id: str) -> tuple[str, str]:
    candidate = Path(model_id)
    if candidate.exists() and candidate.is_dir():
        return str(candidate), "local directory"

    try:
        from huggingface_hub import snapshot_download
    except Exception:
        return model_id, "huggingface (auto-cache)"

    try:
        local = snapshot_download(model_id, local_files_only=True)
        return local, "local cache"
    except Exception:
        local = snapshot_download(model_id)
        return local, "downloaded"

def download_model(model_id: str) -> None:
    if not model_id:
        print("No model id configured (set CUDA_MODEL_ID). Nothing to download.")
        return
    path, source = resolve_model_path(model_id)
    print(f"Model '{model_id}' ready ({source}): {path}")
class Reasoner:
    backend = "none"

    def __init__(self) -> None:
        self.status = "Rule-based mode."

    def is_enabled(self) -> bool:
        return False

    def generate(self, prompt: str, max_new_tokens: int = 180) -> str:
        return ""
class RuleBasedReasoner(Reasoner):
    backend = "rule-based"

    def __init__(self, status: str = "Rule-based mode: no LLM backend enabled.") -> None:
        self.status = status
class CudaReasoner(Reasoner):
    backend = "cuda"

    def __init__(self) -> None:
        self.model: Any = None
        self.tokenizer: Any = None
        self.device = CUDA_DEVICE

        available, detail = detect_cuda()
        if not available:
            self.status = f"CUDA backend unavailable: {detail}"
            return

        if not CUDA_MODEL_ID:
            self.status = f"CUDA detected ({detail}) but CUDA_MODEL_ID is not configured."
            return

        try:
            import torch
            from transformers import (AutoModelForCausalLM,AutoTokenizer)
            model_source, source_label = resolve_model_path(CUDA_MODEL_ID)
            tokenizer: Any = AutoTokenizer.from_pretrained(model_source)
            quant_config: Any = None
            quant_mode = "fp16"
            if CUDA_QUANTIZATION not in ("fp16", "none", "off"):
                try:
                    import bitsandbytes
                    from transformers import BitsAndBytesConfig

                    quant_config = BitsAndBytesConfig(load_in_4bit=True,bnb_4bit_quant_type="nf4",bnb_4bit_compute_dtype=torch.float16,bnb_4bit_use_double_quant=True)
                    quant_mode = "4-bit nf4"
                except Exception:
                    quant_config = None  # bitsandbytes unavailable -> fp16

            if quant_config is not None:
                # 4-bit models are placed via device_map and must not be .to()'d.
                model: Any = AutoModelForCausalLM.from_pretrained(model_source,quantization_config=quant_config,device_map={"": self.device})
            else:
                model = AutoModelForCausalLM.from_pretrained(model_source,torch_dtype=torch.float16)
                model = model.to(self.device)

            model.eval()
            self.tokenizer = tokenizer
            self.model = model
            self.status = (
                            f"CUDA (PyTorch/Transformers, {quant_mode}, {source_label}) loaded: "
                            f"{CUDA_MODEL_ID} on {self.device}. {detail}"
                        )
        except Exception as exc:
            self.model = None
            self.tokenizer = None
            self.status = f"CUDA backend load failed: {exc}"

    def is_enabled(self) -> bool:
        return self.model is not None and self.tokenizer is not None

    def generate(self, prompt: str, max_new_tokens: int = 180) -> str:
        if not self.is_enabled():
            return ""

        try:
            import torch  # type: ignore[import-not-found]

            text = prompt
            if getattr(self.tokenizer, "chat_template", None):
                text = self.tokenizer.apply_chat_template([{"role": "user", "content": prompt}],tokenize=False,add_generation_prompt=True)

            inputs = self.tokenizer(text, return_tensors="pt").to(self.device)
            with torch.no_grad():
                output = self.model.generate(**inputs,max_new_tokens=max_new_tokens,do_sample=False,pad_token_id=self.tokenizer.eos_token_id)
            generated = output[0][inputs["input_ids"].shape[1]:]
            return self.tokenizer.decode(generated, skip_special_tokens=True)
        except Exception:
            return ""
class OpenVINOReasoner(Reasoner):
    backend = "openvino"

    def __init__(self) -> None:
        self.model: Any = None
        self.tokenizer: Any = None
        self.status = "OpenVINO backend unavailable: no model configured (set OPENVINO_MODEL_ID or OPENVINO_MODEL_DIR)."
        if not (OPENVINO_MODEL_DIR or OPENVINO_MODEL_ID):
            return

        try:
            from optimum.intel import OVModelForCausalLM  # type: ignore[import-not-found]
        except Exception as exc:
            self.status = (
                            "OpenVINO backend unavailable: missing dependencies. "
                            f'Install: pip install "optimum[openvino]" nncf  ({exc})'
                            )
            return

        try:
            cache_dir = OV_CACHE_DIR / OPENVINO_MODEL_ID.replace("/", "__")

            if OPENVINO_MODEL_DIR:                              # user-supplied IR dir
                load_dir, export = OPENVINO_MODEL_DIR, False
            elif (cache_dir / "openvino_model.xml").exists():  # previously converted
                load_dir, export = str(cache_dir), False
            else:                                              # first run -> convert + cache
                load_dir, export = OPENVINO_MODEL_ID, True

            tokenizer = self._load_tokenizer(OPENVINO_MODEL_ID if export else load_dir)
            model = self._load_ov_model(OVModelForCausalLM, load_dir, export)

            if export:                                         # persist IR for fast restarts
                try:
                    model.save_pretrained(cache_dir)
                    tokenizer.save_pretrained(cache_dir)
                except Exception:
                    pass

            try:
                model.to(OPENVINO_DEVICE)
            except Exception:
                pass

            self.tokenizer = tokenizer
            self.model = model
            mode = "converted" if export else "cached/local IR"
            self.status = f"OpenVINO (optimum-intel, {mode}) loaded: {OPENVINO_MODEL_ID} on {OPENVINO_DEVICE}."
        except Exception as exc:
            self.model = None
            self.tokenizer = None
            self.status = f"OpenVINO backend load failed: {exc}"

    @staticmethod
    def _load_tokenizer(source: str) -> Any:
        from transformers import AutoTokenizer
        for kwargs in ({"trust_remote_code": True, "fix_mistral_regex": True},
                   {"trust_remote_code": True},
                   {}):
            try:
                return AutoTokenizer.from_pretrained(source, **kwargs)
            except TypeError:
                continue
        return AutoTokenizer.from_pretrained(source)

    @staticmethod
    def _load_ov_model(ov_cls: Any, load_dir: str, export: bool) -> Any:
        base: dict[str, Any] = {"export": export, "trust_remote_code": True}
        attempts: list[dict[str, Any]] = []
        if export and OPENVINO_WEIGHT_FORMAT in ("int8", "8bit"):
            attempts.append({**base, "load_in_8bit": True})
        attempts.append(dict(base))
        attempts.append({k: v for k, v in base.items() if k != "trust_remote_code"})

        last_exc: Exception | None = None
        for kwargs in attempts:
            try:
                return ov_cls.from_pretrained(load_dir, **kwargs)
            except Exception as exc:
                last_exc = exc
        raise last_exc if last_exc else RuntimeError("OpenVINO model load failed.")

    def is_enabled(self) -> bool:
        return self.model is not None and self.tokenizer is not None

    def generate(self, prompt: str, max_new_tokens: int = 180) -> str:
        if not self.is_enabled():
            return ""

        try:
            import torch  # type: ignore[import-not-found]

            text = prompt
            if getattr(self.tokenizer, "chat_template", None):
                text = self.tokenizer.apply_chat_template([{"role": "user", "content": prompt}],tokenize=False,add_generation_prompt=True)

            inputs = self.tokenizer(text, return_tensors="pt")
            with torch.no_grad():
                output = self.model.generate(**inputs,max_new_tokens=max_new_tokens,do_sample=False,pad_token_id=self.tokenizer.eos_token_id)
            generated = output[0][inputs["input_ids"].shape[1]:]
            return self.tokenizer.decode(generated, skip_special_tokens=True)
        except Exception:
            return ""

def build_reasoner() -> Reasoner:
    notes: list[str] = []

    def try_cuda() -> Reasoner | None:
        reasoner = CudaReasoner()
        notes.append(reasoner.status)
        return reasoner if reasoner.is_enabled() else None

    def try_openvino() -> Reasoner | None:
        reasoner = OpenVINOReasoner()
        notes.append(reasoner.status)
        return reasoner if reasoner.is_enabled() else None

    if REASONER_BACKEND == "cuda":
        selected = try_cuda()
    elif REASONER_BACKEND == "openvino":
        selected = try_openvino()
    elif REASONER_BACKEND == "rule":
        selected = None
    else:  # auto
        selected = try_cuda() or try_openvino()

    if selected is not None:
        return selected

    fallback_status = "Rule-based mode (deterministic parsing). " + " | ".join(notes)
    return RuleBasedReasoner(fallback_status.strip())

class PromptAgent:
    def __init__(self, reasoner: Reasoner) -> None:
        self.reasoner = reasoner

    def run(self, prompt: str, location_override: str, trace: Trace) -> dict[str, Any]:
        start = time.perf_counter()
        prompt = prompt.strip()
        location = location_override.strip() or self._extract_location(prompt)
        days = self._extract_days(prompt)

        if self.reasoner.is_enabled() and not location_override.strip():
            llm_location = self._try_llm_extract_location(prompt)
            if llm_location:
                location = llm_location

        if not location:
            raise gr.Error("Location was not detected. Add a location in the prompt or use the Location field.")

        result = {"prompt": prompt, "location": location, "forecast_days": days}
        trace.add("Prompt Agent", "Parse request", "OK", json.dumps(result, ensure_ascii=False), start)
        return result

    def _extract_days(self, prompt: str, default: int = 3) -> int:
        match = re.search(r"(\d+)\s*-?\s*days?\b", prompt, re.IGNORECASE)
        if not match:
            match = re.search(r"\bfor\s+(\d+)\b", prompt, re.IGNORECASE)
        if match:
            try:
                return max(1, min(int(match.group(1)), 30))
            except ValueError:
                pass
        return default

    def _extract_location(self, prompt: str) -> str:
        patterns = [
                    r"(?:weather|forecast)\s+(?:in|for|at)\s+([A-Za-zÁ-ž .'-]{2,60})",
                    r"(?:pocasie|predpoved)\s+(?:v|pre)\s+([A-Za-zÁ-ž .'-]{2,60})",
                    r"(?:in|for|at|v|pre)\s+([A-Za-zÁ-ž .'-]{2,60})",
                    ]
        # Words that signal the location name has ended (the rest is intent,
        # timing, or delivery instructions, not part of the place name).
        stop_words = (
                        r"\b(?:and|then|today|tomorrow|tonight|next|for|please|send|sending|"
                        r"email|e-?mail|mail|as|with|by|pros[ií]m)\b"
                    )
        for pattern in patterns:
            match = re.search(pattern, prompt, re.IGNORECASE)
            if match:
                value = match.group(1).strip(" .,!?:;")
                value = re.split(stop_words, value, flags=re.I)[0]
                return value.strip(" .,!?:;")
        return ""

    def _try_llm_extract_location(self, prompt: str) -> str:
        llm_prompt = (
                        "Extract only the location name from the user request. "
                        "Return JSON only: {\"location\": \"...\"}.\n"
                        f"User request: {prompt}"
                        )
        text = self.reasoner.generate(llm_prompt, max_new_tokens=80)
        try:
            data = json.loads(text[text.find("{") : text.rfind("}") + 1])
            location = str(data.get("location", "")).strip()
            if 2 <= len(location) <= 60:
                return location
        except Exception:
            return ""
        return ""
class WeatherAgent:
    def run(self, request_data: dict[str, Any], trace: Trace) -> dict[str, Any]:
        location = request_data["location"]
        days = request_data["forecast_days"]

        start = time.perf_counter()
        geo = self._geocode(location)
        trace.add("Weather Agent", "Geocode location", "OK", f"{geo['name']}, {geo.get('country', '')}", start)

        start = time.perf_counter()
        forecast = self._forecast(geo, days)
        trace.add("Weather Agent", "Fetch forecast", "OK", f"Fetched {days} day(s) from Open-Meteo", start)

        return {"request": request_data, "geocode": geo, "forecast": forecast}

    def _geocode(self, location: str) -> dict[str, Any]:
        response = requests.get(OPEN_METEO_GEOCODING,params={"name": location, "count": 1, "language": "en", "format": "json"},timeout=20)
        response.raise_for_status()
        data = response.json()
        results = data.get("results") or []
        if not results:
            raise gr.Error(f"Location was not found by Open-Meteo geocoding: {location}")
        return results[0]

    def _forecast(self, geo: dict[str, Any], days: int) -> dict[str, Any]:
        response = requests.get(
            OPEN_METEO_FORECAST,
            params={
                    "latitude": geo["latitude"],
                    "longitude": geo["longitude"],
                    "current": "temperature_2m,relative_humidity_2m,apparent_temperature,precipitation,weather_code,wind_speed_10m",
                    "daily": "weather_code,temperature_2m_max,temperature_2m_min,precipitation_sum,wind_speed_10m_max",
                    "forecast_days": days,
                    "timezone": "auto",
                    },
                    timeout=20,
                    )
        response.raise_for_status()
        return response.json()
class ExcelReportAgent:
    def run(self, weather_payload: dict[str, Any], trace: Trace) -> tuple[Path, pd.DataFrame, dict[str, Any]]:
        start = time.perf_counter()
        geo = weather_payload["geocode"]
        forecast = weather_payload["forecast"]
        request_data = weather_payload["request"]

        daily = forecast.get("daily", {})
        daily_df = pd.DataFrame({
                                "date": daily.get("time", []),
                                "condition": [WEATHER_CODES.get(code, f"Code {code}") for code in daily.get("weather_code", [])],
                                "temp_min_c": daily.get("temperature_2m_min", []),
                                "temp_max_c": daily.get("temperature_2m_max", []),
                                "precipitation_mm": daily.get("precipitation_sum", []),
                                "max_wind_kmh": daily.get("wind_speed_10m_max", []),
                                })

        current = forecast.get("current", {})
        summary = {
                    "requested_location": request_data["location"],
                    "matched_location": f"{geo.get('name')}, {geo.get('country', '')}",
                    "latitude": geo.get("latitude"),
                    "longitude": geo.get("longitude"),
                    "timezone": forecast.get("timezone"),
                    "current_temperature_c": current.get("temperature_2m"),
                    "current_apparent_temperature_c": current.get("apparent_temperature"),
                    "current_humidity_percent": current.get("relative_humidity_2m"),
                    "current_condition": WEATHER_CODES.get(current.get("weather_code"), current.get("weather_code")),
                    "source": "Open-Meteo Forecast API",
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                    }

        safe_location = re.sub(r"[^A-Za-z0-9_-]+", "_", str(geo.get("name", "weather"))).strip("_")
        file_path = OUTPUT_DIR / f"weather_report_{safe_location}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self._write_workbook(file_path, summary, daily_df, trace)

        trace.add("Excel Report Agent", "Create workbook", "OK", str(file_path.name), start)
        return file_path, daily_df, summary

    def _write_workbook(self, file_path: Path, summary: dict[str, Any], daily_df: pd.DataFrame, trace: Trace) -> None:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Summary")
        ws.title = "Summary"
        ws.append(["Field", "Value"])
        for key, value in summary.items():
            ws.append([key, value])

        ws_daily = wb.create_sheet("Daily forecast")
        ws_daily.append(list(daily_df.columns))
        for _, row in daily_df.iterrows():
            ws_daily.append(list(row))

        ws_trace = wb.create_sheet("Agent trace")
        trace_df = trace.dataframe()
        ws_trace.append(list(trace_df.columns))
        for _, row in trace_df.iterrows():
            ws_trace.append(list(row))

        self._style_sheet(ws)
        self._style_sheet(ws_daily)
        self._style_sheet(ws_trace)
        wb.save(file_path)

    def _style_sheet(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)

        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, ws.max_row + 1))
            ws.column_dimensions[column_letter].width = min(max(max_len + 3, 14), 42)

        ws.freeze_panes = "A2"
class EmailAgent:
    EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

    def _parse_recipients(self, raw: str) -> tuple[list[str], list[str]]:
        parts = [p.strip() for p in re.split(r"[,;\n]+", raw or "")]
        seen: set[str] = set()
        valid: list[str] = []
        invalid: list[str] = []
        for part in parts:
            if not part or part.lower() in seen:
                continue
            seen.add(part.lower())
            (valid if self.EMAIL_RE.match(part) else invalid).append(part)
        return valid, invalid

    def run(self,file_path: Path,summary: dict[str, Any],should_send: bool,recipient: str,trace: Trace) -> str:
        start = time.perf_counter()
        recipient = (recipient or "").strip() or PREDEFINED_RECIPIENT

        if not should_send:
            trace.add("Email Agent", "Send email", "SKIPPED", "Dry run: send checkbox not enabled", start)
            return "Email Agent skipped: dry run mode."

        recipients, invalid = self._parse_recipients(recipient)

        if invalid:
            detail = "Invalid recipient email address(es): " + ", ".join(invalid)
            trace.add("Email Agent", "Send email", "SKIPPED", detail, start)
            return detail

        if not recipients:
            detail = "No recipient email provided. Enter one or more addresses (comma-separated) or set PREDEFINED_RECIPIENT."
            trace.add("Email Agent", "Send email", "SKIPPED", detail, start)
            return detail

        missing = [
                    name
                    for name, value in {"SMTP_HOST": SMTP_HOST,"SMTP_FROM": SMTP_FROM}.items()
                    if not value
                    ]
        if missing:
            detail = "Missing email configuration: " + ", ".join(missing)
            trace.add("Email Agent", "Send email", "SKIPPED", detail, start)
            return detail

        msg = EmailMessage()
        msg["Subject"] = f"Weather report - {summary.get('matched_location', 'location')}"
        msg["From"] = SMTP_FROM
        msg["To"] = ", ".join(recipients)
        msg.set_content(
                        "Hello,\n\n"
                        "I have attached the weather report generated by the Agentic AI demo pipeline.\n\n"
                        f"Location: {summary.get('matched_location')}\n"
                        f"Current condition: {summary.get('current_condition')}\n"
                        f"Current temperature: {summary.get('current_temperature_c')} °C\n\n"
                        "Regards, \n Agentic AI Weather Demo"
                        )

        with file_path.open("rb") as fh:
                msg.add_attachment(
                                    fh.read(),
                                    maintype="application",
                                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    filename=file_path.name,
                                    )

        context = ssl.create_default_context()
        use_implicit_ssl = SMTP_SECURITY in ("ssl", "smtps", "tls")
        use_starttls = SMTP_SECURITY not in ("ssl", "smtps", "tls", "none", "plain", "off")

        try:
            server: smtplib.SMTP
            if use_implicit_ssl:
                server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context, timeout=30)
            else:
                server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)

            with server:
                if use_starttls:
                    server.starttls(context=context)
                if SMTP_USER and SMTP_PASSWORD:
                    server.login(SMTP_USER, SMTP_PASSWORD)
                server.send_message(msg)
        except (OSError, smtplib.SMTPException) as exc:
            detail = (
                        f"Email send failed via {SMTP_HOST}:{SMTP_PORT} ({SMTP_SECURITY}): {exc}. "
                        "Check SMTP_HOST/SMTP_PORT, network access, and credentials."
                    )
            trace.add("Email Agent", "Send email", "ERROR", detail, start)
            return detail

        security_label = "implicit SSL" if use_implicit_ssl else ("STARTTLS" if use_starttls else "no encryption")
        trace.add("Email Agent", "Send email", "OK", f"Sent to: {recipient} ({security_label})", start)
        return f"Email sent to: {recipient}"

REASONER = build_reasoner()
PROMPT_AGENT = PromptAgent(REASONER)
WEATHER_AGENT = WeatherAgent()
EXCEL_AGENT = ExcelReportAgent()
EMAIL_AGENT = EmailAgent()

def run_pipeline(prompt: str,location_override: str,send_email: bool,recipient_email: str):
    trace = Trace()
    started = time.perf_counter()

    request_data = PROMPT_AGENT.run(prompt, location_override, trace)
    weather_payload = WEATHER_AGENT.run(request_data, trace)
    file_path, daily_df, summary = EXCEL_AGENT.run(weather_payload, trace)
    email_status = EMAIL_AGENT.run(file_path, summary, bool(send_email), recipient_email, trace)

    total_sec = round(time.perf_counter() - started, 3)
    status = (
                f"Pipeline finished in {total_sec}s\n\n"
                f"Reasoning backend: {REASONER.backend}\n"
                f"Runtime: {REASONER.status}\n"
                f"Weather source: Open-Meteo\n"
                f"Forecast days (from prompt): {request_data['forecast_days']}\n"
                f"Excel file: {file_path.name}\n"
                f"Email status: {email_status}\n\n"
                f"Current weather in {summary.get('matched_location')}: "
                f"{summary.get('current_condition')}, {summary.get('current_temperature_c')} °C"
                )

    return status, trace.dataframe(), daily_df, str(file_path)


def _launch_supports_css() -> bool:
    try:
        import inspect

        return "css" in inspect.signature(gr.Blocks.launch).parameters
    except Exception:
        return False

def build_ui() -> gr.Blocks:
    blocks_kwargs: dict[str, Any] = {"title": "Agentic AI Weather Pipeline"}
    if not _launch_supports_css():
        blocks_kwargs["css"] = CUSTOM_CSS
    with gr.Blocks(**blocks_kwargs) as demo:
        gr.Markdown(
                    """
                    # Agentic AI Weather Pipeline

                    **Prompt Agent -> Weather Agent -> Excel Agent -> Email Agent**  
                    Local OpenVINO reasoning is optional. The workflow remains deterministic and auditable.
                    """
                    )

        with gr.Row():
            with gr.Column(scale=1):
                prompt = gr.Textbox(
                                    label="User prompt",
                                    lines=5,
                                    value="Please prepare a 3-day weather report for Bratislava and send it as Excel.",
                                    info="Forecast length is read from the text, e.g. '3-day' or 'for 5 days' (1-10, default 3).",
                                    elem_classes=["agent-box"],
                                    )
                location = gr.Textbox(
                                        label="Location override (optional)",
                                        value="",
                                        placeholder="e.g. Bratislava, Vienna, Prague",
                                    )
                recipient = gr.Textbox(
                                        label="Recipient email",
                                        value=PREDEFINED_RECIPIENT,
                                        lines=3,
                                        placeholder="name@example.com",
                                        info="Defaults to PREDEFINED_RECIPIENT; edit to send elsewhere. Validated before sending.",
                                        )
                send = gr.Checkbox(False, label="Send email to the recipient above")
                run_btn = gr.Button("Run agent pipeline", variant="primary")

            with gr.Column(scale=1):
                status = gr.Textbox(label="Pipeline result", lines=15, elem_classes=["agent-box"])
                excel_file = gr.File(label="Generated Excel report")

        with gr.Row():
            trace = gr.Dataframe(label="Agent trace", interactive=False)
            daily = gr.Dataframe(label="Daily forecast", interactive=False)

        gr.Markdown(
                    f"""
                    ### Runtime status

                    `{REASONER.status}`

                    The recipient email defaults to `PREDEFINED_RECIPIENT` but can be edited in the interface. The address is validated before sending, and SMTP credentials still come from environment variables.
                    """
                    )

        run_btn.click(run_pipeline,[prompt, location, send, recipient],[status, trace, daily, excel_file])

    return demo

if __name__ == "__main__":
    import sys
    if "--download" in sys.argv:
        download_model(CUDA_MODEL_ID)
        sys.exit(0)

    app = build_ui()
    launch_kwargs: dict[str, Any] = dict(
                                        server_name=os.getenv("GRADIO_SERVER_NAME", "127.0.0.1"),
                                        #server_port=int(os.getenv("GRADIO_SERVER_PORT", "7861")),
                                        inbrowser=True,
                                        share=False,
                                        allowed_paths=[str(OUTPUT_DIR)],
                                        )
    if _launch_supports_css():
        launch_kwargs["css"] = CUSTOM_CSS
    app.launch(**launch_kwargs)